from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from rest_framework.decorators import action
from django.contrib.auth import get_user_model
from .models import Audit, NonConformite, GridRow
from .serializers import AuditSerializer, NonConformiteSerializer, UserSerializer, GridRowSerializer
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.authentication import TokenAuthentication
import pandas as pd
import os
from django.conf import settings
import threading
import json
import csv
from django.utils.timezone import make_aware
from datetime import datetime, timedelta
from pytz import timezone

class RegisterView(APIView):
    def post(self, request):
        email = request.data.get('email', '')
        if not email.endswith('@stellantis.com'):
            return Response({'error': 'Email must be @stellantis.com'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            user = get_user_model().objects.create_user(
                username=request.data['username'],
                email=email,
                password=request.data['password'],
                role=request.data.get('role', 'operator'),
                shift=request.data.get('shift', None)
            )
            return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LoginView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if user:
            token, _ = Token.objects.get_or_create(user=user)
            return Response({'token': token.key, 'user': UserSerializer(user).data})
        # Try email login
        try:
            user_model = get_user_model()
            user_obj = user_model.objects.get(email=username)
            user = authenticate(username=user_obj.username, password=password)
            if user:
                token, _ = Token.objects.get_or_create(user=user)
                return Response({'token': token.key, 'user': UserSerializer(user).data})
        except Exception:
            pass
        return Response({'error': 'Invalid credentials'}, status=status.HTTP_400_BAD_REQUEST)

class AuditViewSet(viewsets.ModelViewSet):
    queryset = Audit.objects.all().order_by('-date')
    serializer_class = AuditSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def create(self, request, *args, **kwargs):
        # Prevent supervisors from creating audits
        if hasattr(request.user, 'role') and request.user.role == 'supervisor':
            return Response({'detail': 'Supervisors are not allowed to submit audits.'}, status=status.HTTP_403_FORBIDDEN)
        # Always work with a mutable dict
        data = request.data.copy()
        files = request.FILES

        # 1. Parse nonconformites JSON string
        nonconformites_raw = data.get('nonconformites')
        if not nonconformites_raw:
            return Response({'nonconformites': ['Ce champ est obligatoire.']}, status=status.HTTP_400_BAD_REQUEST)
        try:
            if isinstance(nonconformites_raw, str):
                nonconformites = json.loads(nonconformites_raw)
            else:
                nonconformites = nonconformites_raw
        except Exception:
            return Response({'nonconformites': ['Format JSON invalide.']}, status=status.HTTP_400_BAD_REQUEST)

        # 2. Handle nested list (array in array)
        if isinstance(nonconformites, list) and len(nonconformites) == 1 and isinstance(nonconformites[0], list):
            nonconformites = nonconformites[0]

        # 3. Replace photo_field with actual file
        for nc in nonconformites:
            photo_field = nc.pop('photo_field', None)
            if photo_field and photo_field in files:
                nc['photo'] = files[photo_field]

        # 4. Build a plain dict for the serializer (not a QueryDict)
        # Fix: ensure all values are strings (not lists) for simple fields
        payload = {}
        for key, value in data.items():
            if key == 'nonconformites':
                continue
            if isinstance(value, list):
                payload[key] = value[0] if value else ''
            else:
                payload[key] = value
        payload['nonconformites'] = nonconformites

        serializer = self.get_serializer(data=payload)
        if not serializer.is_valid():
            print('AUDIT SUBMISSION ERROR:', serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class NonConformiteViewSet(viewsets.ModelViewSet):
    queryset = NonConformite.objects.all()
    serializer_class = NonConformiteSerializer
    permission_classes = [IsAuthenticated]

grid_schema_cache = {
    'data': None,
    'lock': threading.Lock()
}

def load_grid_schema():
    with grid_schema_cache['lock']:
        if grid_schema_cache['data'] is None:
            excel_path = os.path.join(settings.BASE_DIR, 'grille_audit_stellantis.xlsx')
            df = pd.read_excel(excel_path)
            # Normalize columns for frontend compatibility
            col_map = {
                'Catégorie': 'category',
                'Categorie': 'category',
                'category': 'category',
                'Libellé': 'label',
                'Libelle': 'label',
                'label': 'label',
                'Code anomalie': 'code_anomalie',
                'Code': 'code_anomalie',
                'code_anomalie': 'code_anomalie',
                'Chapitre MLP': 'chapitre_mlp',
                'Chapitre': 'chapitre_mlp',
                'chapitre_mlp': 'chapitre_mlp',
            }
            df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
            grid_schema_cache['data'] = json.loads(df.to_json(orient='records', force_ascii=False))
    return grid_schema_cache['data']

class GridSchemaView(APIView):
    permission_classes = [permissions.AllowAny]
    def get(self, request):
        rows = GridRow.objects.all()
        data = GridRowSerializer(rows, many=True).data
        return Response(data)

class AuditExportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, pk, format=None):
        print(f"Export requested for audit ID: {pk}")
        try:
            audit = Audit.objects.get(pk=pk)
        except Audit.DoesNotExist:
            print("Audit not found!")
            return Response({'error': 'Audit not found'}, status=404)
        user = request.user
        # Permission logic: supervisor can export any, operator only their own
        if hasattr(user, 'role'):
            if user.role == 'supervisor':
                pass  # allow
            elif user.role == 'operator' and audit.user_id != user.id:
                return Response({'detail': 'Vous n\'avez pas accès à cet audit.'}, status=403)
        else:
            # fallback: only allow owner
            if audit.user_id != user.id:
                return Response({'detail': 'Vous n\'avez pas accès à cet audit.'}, status=403)
        export_format = request.GET.get('format', 'csv')
        if export_format == 'csv':
            import io
            import pandas as pd
            # Build DataFrame for the grid
            rows = []
            for nc in audit.nonconformites.all():
                rows.append({
                    'Catégorie': nc.category,
                    'Libellé': nc.label,
                    'Code anomalie': nc.code_anomalie,
                    'Chapitre MLP': nc.chapitre_mlp,
                    'UM': nc.um,
                    'UC': nc.uc,
                    'UGS': nc.ugs,
                    'AVEXP': nc.avexp,
                    'Remarque': nc.remark,
                })
            df = pd.DataFrame(rows)
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename=grille_audit_stellantis_{pk}.csv'
            return response
        return Response({'error': 'Format not supported'}, status=400)

class AuditStatsView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.user
        date_str = request.GET.get('date')
        shift = request.GET.get('shift')
        # Convert Morocco local date range to UTC for filtering
        from pytz import timezone as pytz_timezone
        tz = pytz_timezone('Africa/Casablanca')
        utc = pytz_timezone('UTC')
        if date_str:
            # Always interpret date_str as Morocco local time
            local_start = tz.localize(datetime.strptime(date_str, '%Y-%m-%d'))
        else:
            now = datetime.now(tz)
            local_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        local_end = local_start + timedelta(days=1)
        # Convert local Morocco time range to UTC for filtering
        day_start = local_start.astimezone(utc)
        day_end = local_end.astimezone(utc)
        print(f"AuditStatsView: day_start={day_start}, day_end={day_end}, tz=UTC (filtering Morocco local day)")
        if user.role == 'supervisor':
            audits = Audit.objects.filter(date__gte=day_start, date__lt=day_end)
            if shift and shift != 'All':
                audits = audits.filter(user__shift=shift)
        else:
            audits = Audit.objects.filter(user=user, date__gte=day_start, date__lt=day_end)
        print(f"AuditStatsView: audits count={audits.count()}")
        for audit in audits:
            print(f"Audit id={audit.id}, date={audit.date}")
        total = audits.count()
        defects = sum(nc.um or nc.uc or nc.ugs or nc.avexp for audit in audits for nc in audit.nonconformites.all())
        serializer = AuditSerializer(audits, many=True)
        return Response({
            'total': total,
            'defects': defects,
            'audits': serializer.data
        })

class ExportAuditCSVView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        import traceback
        try:
            print(f"EXPORT AUDIT CSV CLASS VIEW CALLED for pk={pk}, user={request.user} (id={request.user.id}, role={getattr(request.user, 'role', None)})")
            # Try both int and str PKs for robustness
            audit = None
            try:
                audit = Audit.objects.get(pk=pk)
            except Audit.DoesNotExist:
                try:
                    audit = Audit.objects.get(pk=int(pk))
                except (Audit.DoesNotExist, ValueError):
                    print(f"Audit with pk={pk} not found.")
                    from django.http import JsonResponse
                    return JsonResponse({'error': f'Audit {pk} not found'}, status=404)
            user = request.user
            if hasattr(user, 'role'):
                if user.role == 'supervisor':
                    print(f"Supervisor access granted for audit {pk}.")
                elif user.role == 'operator' and audit.user_id != user.id:
                    print(f"Operator {user.id} forbidden from exporting audit {pk} owned by {audit.user_id}.")
                    from django.http import JsonResponse
                    return JsonResponse({'detail': 'Vous n\'avez pas accès à cet audit.'}, status=403)
            else:
                if audit.user_id != user.id:
                    print(f"User {user.id} forbidden from exporting audit {pk} owned by {audit.user_id}.")
                    from django.http import JsonResponse
                    return JsonResponse({'detail': 'Vous n\'avez pas accès à cet audit.'}, status=403)
            print(f"Exporting audit {pk} for user {user.id}.")

            # Build Excel file using openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Stellantis"

            # Define headers and columns
            headers = [
                'Catégorie', 'Libellé', 'Code anomalie', 'Chapitre MLP', 'UM', 'UC', 'UGS', 'AVEXP', 'Remarque', 'Conforme', 'Non Conforme'
            ]
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Add rows
            for nc in audit.nonconformites.all():
                row = [
                    nc.category,
                    nc.label,
                    nc.code_anomalie,
                    nc.chapitre_mlp,
                    nc.um,
                    nc.uc,
                    nc.ugs,
                    nc.avexp,
                    nc.remark,
                    "",  # Conforme checkbox
                    ""   # Non Conforme checkbox
                ]
                ws.append(row)

            # Add checkboxes (data validation: TRUE/FALSE dropdown)
            dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
            ws.add_data_validation(dv)
            for row in range(2, ws.max_row + 1):
                dv.add(ws.cell(row=row, column=10))  # Conforme
                dv.add(ws.cell(row=row, column=11))  # Non Conforme

            # Auto-size columns with a minimum width to avoid '###' in Excel
            MIN_WIDTH = 12
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column].width = max(max_length + 2, MIN_WIDTH)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            from django.http import HttpResponse
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=grille_audit_stellantis_{pk}.xlsx'
            return response
        except Exception as e:
            print("[EXPORT ERROR] Exception occurred:")
            traceback.print_exc()
            from django.http import JsonResponse
            return JsonResponse({'error': 'Une erreur est survenue lors de l\'exportation.'}, status=500)
